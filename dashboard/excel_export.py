"""
Utilitaires pour l'export Excel des tableaux de bord
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from .models import Objectives, Indicateur, Cible, Periodicite, Observation
import logging

logger = logging.getLogger(__name__)


def export_tableau_bord_to_excel(tableau_bord):
    """
    Exporte un tableau de bord vers un fichier Excel
    
    Structure du tableau:
    - Colonnes: N°, Objectifs, Indicateurs, Fréquence, Cible, 
                T1 (A réaliser, Réalisé, Taux), T2 (...), T3 (...), T4 (...), Observations
    - Total: 18 colonnes
    """
    
    # Créer un nouveau workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Tableau de Bord"
    
    # Styles
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    subheader_font = Font(name='Calibri', size=10, bold=True, color='000000')
    subheader_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    subheader_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    cell_alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # En-tête du document
    ws.merge_cells('A1:R1')
    title_cell = ws['A1']
    title_cell.value = f"Tableau de Bord - {tableau_bord.processus.nom_processus if tableau_bord.processus else 'N/A'}"
    title_cell.font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    title_cell.fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25
    
    # Informations du tableau
    ws.merge_cells('A2:R2')
    info_cell = ws['A2']
    info_cell.value = f"Année: {tableau_bord.annee} | Type: {tableau_bord.type_tableau.libelle if tableau_bord.type_tableau else 'N/A'}"
    info_cell.font = Font(name='Calibri', size=10, italic=True)
    info_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 20
    
    # En-têtes de premier niveau (ligne 3)
    headers_level1 = [
        ('A3', 'B3', 'Informations générales'),
        ('C3', 'E3', 'Indicateurs et mesures'),
        ('F3', 'H3', 'Trimestre 1'),
        ('I3', 'K3', 'Trimestre 2'),
        ('L3', 'N3', 'Trimestre 3'),
        ('O3', 'Q3', 'Trimestre 4'),
        ('R3', 'R3', 'Observations')
    ]
    
    for start, end, text in headers_level1:
        ws.merge_cells(f'{start}:{end}')
        cell = ws[start]
        cell.value = text
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    ws.row_dimensions[3].height = 30
    
    # En-têtes de second niveau (ligne 4)
    headers_level2 = [
        ('A4', 'N°'),
        ('B4', 'OBJECTIFS'),
        ('C4', 'INDICATEURS'),
        ('D4', 'FRÉQUENCE'),
        ('E4', 'CIBLE'),
        ('F4', 'A réaliser'),
        ('G4', 'Réalisé'),
        ('H4', 'Taux'),
        ('I4', 'A réaliser'),
        ('J4', 'Réalisé'),
        ('K4', 'Taux'),
        ('L4', 'A réaliser'),
        ('M4', 'Réalisé'),
        ('N4', 'Taux'),
        ('O4', 'A réaliser'),
        ('P4', 'Réalisé'),
        ('Q4', 'Taux'),
        ('R4', 'OBSERVATIONS')
    ]
    
    for cell_ref, text in headers_level2:
        cell = ws[cell_ref]
        cell.value = text
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = subheader_alignment
        cell.border = thin_border
    
    ws.row_dimensions[4].height = 40
    
    # Largeurs des colonnes
    column_widths = {
        'A': 8,   # N°
        'B': 40,  # Objectifs
        'C': 45,  # Indicateurs
        'D': 15,  # Fréquence
        'E': 12,  # Cible
        'F': 12,  # T1 A réaliser
        'G': 12,  # T1 Réalisé
        'H': 10,  # T1 Taux
        'I': 12,  # T2 A réaliser
        'J': 12,  # T2 Réalisé
        'K': 10,  # T2 Taux
        'L': 12,  # T3 A réaliser
        'M': 12,  # T3 Réalisé
        'N': 10,  # T3 Taux
        'O': 12,  # T4 A réaliser
        'P': 12,  # T4 Réalisé
        'Q': 10,  # T4 Taux
        'R': 30   # Observations
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Récupérer les données
    objectives = Objectives.objects.filter(tableau_bord=tableau_bord).order_by('number')
    
    # Remplir les données (à partir de la ligne 5)
    current_row = 5
    
    for objective in objectives:
        # Récupérer les indicateurs de cet objectif
        indicateurs = Indicateur.objects.filter(objective_id=objective.uuid).order_by('created_at')
        
        if indicateurs.exists():
            # Calculer la hauteur de ligne nécessaire
            num_indicateurs = indicateurs.count()
            row_height = max(30, num_indicateurs * 25)
            ws.row_dimensions[current_row].height = row_height
            
            # N° Objectif
            cell = ws[f'A{current_row}']
            cell.value = objective.number
            cell.font = Font(name='Calibri', size=10, bold=True)
            cell.alignment = center_alignment
            cell.border = thin_border
            
            # Libellé Objectif
            cell = ws[f'B{current_row}']
            cell.value = objective.libelle
            cell.font = Font(name='Calibri', size=10)
            cell.alignment = cell_alignment
            cell.border = thin_border
            
            # Indicateurs (colonne C)
            indicateurs_text = '\n\n'.join([ind.libelle for ind in indicateurs])
            cell = ws[f'C{current_row}']
            cell.value = indicateurs_text
            cell.font = Font(name='Calibri', size=10)
            cell.alignment = cell_alignment
            cell.border = thin_border
            
            # Fréquence (colonne D)
            frequences_text = '\n\n'.join([
                ind.frequence_id.nom if ind.frequence_id else 'Non définie'
                for ind in indicateurs
            ])
            cell = ws[f'D{current_row}']
            cell.value = frequences_text
            cell.font = Font(name='Calibri', size=9)
            cell.alignment = center_alignment
            cell.border = thin_border
            
            # Cible (colonne E)
            cibles_text = []
            for ind in indicateurs:
                try:
                    cible = Cible.objects.get(indicateur_id=ind.uuid)
                    cibles_text.append(f"{cible.condition} {cible.valeur}")
                except Cible.DoesNotExist:
                    cibles_text.append('-')
            
            cell = ws[f'E{current_row}']
            cell.value = '\n\n'.join(cibles_text)
            cell.font = Font(name='Calibri', size=9)
            cell.alignment = center_alignment
            cell.border = thin_border
            
            # Périodicités pour chaque trimestre (colonnes F-Q)
            periodes = ['T1', 'T2', 'T3', 'T4']
            col_start = ['F', 'I', 'L', 'O']
            
            for periode_idx, periode in enumerate(periodes):
                col_base = col_start[periode_idx]
                
                # A réaliser
                a_realiser_values = []
                for ind in indicateurs:
                    try:
                        periodicite = Periodicite.objects.get(indicateur_id=ind.uuid, periode=periode)
                        a_realiser_values.append(str(periodicite.a_realiser) if periodicite.a_realiser else '-')
                    except Periodicite.DoesNotExist:
                        a_realiser_values.append('-')
                
                cell = ws[f'{col_base}{current_row}']
                cell.value = '\n\n'.join(a_realiser_values)
                cell.font = Font(name='Calibri', size=9)
                cell.alignment = center_alignment
                cell.border = thin_border
                
                # Réalisé
                col_realise = chr(ord(col_base) + 1)
                realise_values = []
                for ind in indicateurs:
                    try:
                        periodicite = Periodicite.objects.get(indicateur_id=ind.uuid, periode=periode)
                        realise_values.append(str(periodicite.realiser) if periodicite.realiser else '-')
                    except Periodicite.DoesNotExist:
                        realise_values.append('-')
                
                cell = ws[f'{col_realise}{current_row}']
                cell.value = '\n\n'.join(realise_values)
                cell.font = Font(name='Calibri', size=9)
                cell.alignment = center_alignment
                cell.border = thin_border
                
                # Taux
                col_taux = chr(ord(col_base) + 2)
                taux_values = []
                for ind in indicateurs:
                    try:
                        periodicite = Periodicite.objects.get(indicateur_id=ind.uuid, periode=periode)
                        if periodicite.a_realiser and periodicite.realiser and periodicite.a_realiser > 0:
                            taux = (periodicite.realiser / periodicite.a_realiser) * 100
                            taux_values.append(f"{taux:.1f}%")
                        else:
                            taux_values.append('-')
                    except (Periodicite.DoesNotExist, ZeroDivisionError, TypeError):
                        taux_values.append('-')
                
                cell = ws[f'{col_taux}{current_row}']
                cell.value = '\n\n'.join(taux_values)
                cell.font = Font(name='Calibri', size=9)
                cell.alignment = center_alignment
                cell.border = thin_border
            
            # Observations (colonne R)
            observations_text = []
            for ind in indicateurs:
                try:
                    observation = Observation.objects.get(indicateur_id=ind.uuid)
                    observations_text.append(observation.libelle)
                except Observation.DoesNotExist:
                    observations_text.append('-')
            
            cell = ws[f'R{current_row}']
            cell.value = '\n\n'.join(observations_text)
            cell.font = Font(name='Calibri', size=9)
            cell.alignment = cell_alignment
            cell.border = thin_border
            
        else:
            # Pas d'indicateurs pour cet objectif
            ws.row_dimensions[current_row].height = 25
            
            # N° Objectif
            cell = ws[f'A{current_row}']
            cell.value = objective.number
            cell.font = Font(name='Calibri', size=10, bold=True)
            cell.alignment = center_alignment
            cell.border = thin_border
            
            # Libellé Objectif
            cell = ws[f'B{current_row}']
            cell.value = objective.libelle
            cell.font = Font(name='Calibri', size=10)
            cell.alignment = cell_alignment
            cell.border = thin_border
            
            # Remplir les autres colonnes avec "-"
            for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R']:
                cell = ws[f'{col}{current_row}']
                cell.value = '-'
                cell.font = Font(name='Calibri', size=9, color='999999')
                cell.alignment = center_alignment
                cell.border = thin_border
        
        current_row += 1
    
    # Si aucun objectif
    if not objectives.exists():
        cell = ws['A5']
        ws.merge_cells('A5:R5')
        cell.value = "Aucun objectif défini pour ce tableau de bord"
        cell.font = Font(name='Calibri', size=11, italic=True, color='999999')
        cell.alignment = center_alignment
        ws.row_dimensions[5].height = 40
    
    return wb


def generate_excel_response(tableau_bord):
    """
    Génère une réponse HTTP avec le fichier Excel
    """
    wb = export_tableau_bord_to_excel(tableau_bord)
    
    # Créer la réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Nom du fichier
    filename = f"Tableau_Bord_{tableau_bord.processus.nom_processus if tableau_bord.processus else 'Export'}_{tableau_bord.annee}.xlsx"
    filename = filename.replace(' ', '_').replace('/', '_')
    
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Sauvegarder le workbook dans la réponse
    wb.save(response)
    
    return response
